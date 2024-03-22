import os
import re
from docx import Document
from openpyxl import load_workbook
from PyPDF2 import PdfReader

def extract_metadata_from_pdf(filename):
    pdf_file = PdfReader(filename)
    metadata = {
        "author": pdf_file.metadata.author,
        "creator": pdf_file.metadata.creator,
        "producer": pdf_file.metadata.producer,
        "title": pdf_file.metadata.title,
        "subject": pdf_file.metadata.subject,
        "created": pdf_file.metadata.get('/CreationDate', None),
        "modified": pdf_file.metadata.get('/ModDate', None)
    }
    return metadata

def extract_metadata_from_docx(filename):
    document = Document(filename)
    core_properties = document.core_properties
    metadata = {
        "author": core_properties.author,
        "title": core_properties.title,
        "created": core_properties.created,
        "modified": core_properties.modified
    }
    return metadata

def extract_metadata_from_xlsx(filename):
    wb = load_workbook(filename)
    author = None
    if wb.sheetnames:  # Verifica si hay hojas de cálculo en el libro
        sheet = wb[wb.sheetnames[0]]  # Obtiene la primera hoja de cálculo
        if sheet.dimensions:  # Verifica si la hoja de cálculo contiene datos
            author = sheet.cell(row=1, column=1).value  # Obtiene el valor de la primera celda
    metadata = {
        "author": author,
        "title": wb.properties.title,
        "created": wb.properties.created,
        "modified": wb.properties.modified
    }
    return metadata

def process_files(directory_path):
    for file_name in os.listdir(directory_path):
        file_path = os.path.join(directory_path, file_name)
        if os.path.isfile(file_path):
            file_extension = re.findall(r"\.(pdf|docx|xlsx)$", file_name)
            if file_extension:
                file_extension = file_extension[0]
                if file_extension == "pdf":
                    metadata = extract_metadata_from_pdf(file_path)
                elif file_extension == "docx":
                    metadata = extract_metadata_from_docx(file_path)
                elif file_extension == "xlsx":
                    metadata = extract_metadata_from_xlsx(file_path)
                print(f"Metadatos del archivo {file_name}:")
                for k, v in metadata.items():
                    print(f"  {k}: {v}")
                print()

# Ruta al directorio
directory_path = "/Users/aldoalcantara/Documents/ESCOM/Computer Security/Python/All_Meta"

# Procesar los archivos en el directorio
process_files(directory_path)